"""Tests for the Excel (.xlsx) parser."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from pith.parsers.base import ParseError
from pith.parsers.xlsx import parse_xlsx

FIXTURES = Path(__file__).parent.parent / "fixtures"


@pytest.fixture(autouse=True)
def _ensure_fixtures() -> None:
    FIXTURES.mkdir(parents=True, exist_ok=True)


def _make_workbook(path: Path, sheets: dict[str, list[list[str]]]) -> Path:
    """Create an .xlsx file with the given sheet data."""
    wb = Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    wb.save(str(path))
    return path


class TestParseXlsx:
    def test_single_sheet(self, tmp_path: Path) -> None:
        path = _make_workbook(
            tmp_path / "single.xlsx",
            {"Data": [["Name", "Age"], ["Alice", "30"], ["Bob", "25"]]},
        )
        doc = parse_xlsx(path)

        assert doc.title == "single"
        assert len(doc.sections) == 1
        assert doc.sections[0].heading == "Data"
        assert len(doc.tables) == 1
        assert doc.tables[0].headers == ["Name", "Age"]
        assert doc.tables[0].rows == [["Alice", "30"], ["Bob", "25"]]
        assert doc.metadata["source"] == str(path)
        assert doc.metadata["format"] == "xlsx"
        assert doc.metadata["sheet_count"] == "1"
        assert doc.metadata["row_count_Data"] == "2"

    def test_multiple_sheets(self, tmp_path: Path) -> None:
        path = _make_workbook(
            tmp_path / "multi.xlsx",
            {
                "Sheet1": [["X", "Y"], ["1", "2"]],
                "Sheet2": [["A", "B"], ["3", "4"], ["5", "6"]],
            },
        )
        doc = parse_xlsx(path)

        assert len(doc.sections) == 2
        assert len(doc.tables) == 2
        assert doc.sections[0].heading == "Sheet1"
        assert doc.sections[1].heading == "Sheet2"
        assert doc.tables[1].rows == [["3", "4"], ["5", "6"]]
        assert doc.metadata["sheet_count"] == "2"
        assert doc.metadata["row_count_Sheet1"] == "1"
        assert doc.metadata["row_count_Sheet2"] == "2"

    def test_empty_sheets_skipped(self, tmp_path: Path) -> None:
        path = tmp_path / "with_empty.xlsx"
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Empty"
        ws2 = wb.create_sheet(title="HasData")
        ws2.append(["Col1"])
        ws2.append(["Val1"])
        wb.save(str(path))

        doc = parse_xlsx(path)

        assert len(doc.sections) == 1
        assert doc.sections[0].heading == "HasData"

    def test_all_empty_sheets_raises(self, tmp_path: Path) -> None:
        path = tmp_path / "all_empty.xlsx"
        wb = Workbook()
        wb.active.title = "Empty"
        wb.save(str(path))

        with pytest.raises(ParseError) as exc_info:
            parse_xlsx(path)
        assert exc_info.value.path == path

    def test_corrupt_file_raises(self, tmp_path: Path) -> None:
        path = tmp_path / "corrupt.xlsx"
        path.write_bytes(b"not an xlsx file")

        with pytest.raises(ParseError) as exc_info:
            parse_xlsx(path)
        assert exc_info.value.path == path

    def test_none_cells_become_empty_string(self, tmp_path: Path) -> None:
        path = tmp_path / "nulls.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sparse"
        ws.append(["A", "B", "C"])
        ws.append([1, None, 3])
        wb.save(str(path))

        doc = parse_xlsx(path)

        assert doc.tables[0].rows[0] == ["1", "", "3"]

    def test_dispatch_integration(self, tmp_path: Path) -> None:
        """Verify .xlsx routes through the dispatch map."""
        from pith.parsers import parse

        path = _make_workbook(
            tmp_path / "dispatch.xlsx",
            {"Sheet1": [["H1"], ["V1"]]},
        )
        doc = parse(path)
        assert doc.metadata["format"] == "xlsx"
