"""Excel parser — extraction via openpyxl.

Each sheet becomes a Section (heading = sheet title).
Each sheet's data becomes a Table (first row = headers).
Empty sheets are skipped silently.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from pith.i18n import t
from pith.parsers.base import ParsedDocument, ParseError, Section, Table


def parse_xlsx(path: Path) -> ParsedDocument:
    """Parse an Excel file into a ``ParsedDocument``.

    Each non-empty sheet produces one Section (heading = sheet title)
    and one Table (first row = headers, remaining rows = data).

    Args:
        path: Path to the ``.xlsx`` file.

    Returns:
        A ``ParsedDocument`` with one section and table per sheet.

    Raises:
        ParseError: If the file cannot be read or is corrupt.
    """
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except (InvalidFileException, Exception) as exc:
        raise ParseError(path, t("parser.xlsx_read_error", detail=str(exc))) from exc

    sections: list[Section] = []
    tables: list[Table] = []
    sheet_row_counts: dict[str, int] = {}

    try:
        for ws in wb.worksheets:
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                cell_values = [str(v) if v is not None else "" for v in row]
                if any(v != "" for v in cell_values):
                    rows.append(cell_values)

            if not rows:
                continue

            headers = rows[0]
            data_rows = rows[1:]
            sheet_row_counts[ws.title] = len(data_rows)

            body_lines = ["\t".join(r) for r in rows]
            sections.append(Section(heading=ws.title, body="\n".join(body_lines)))
            tables.append(Table(headers=headers, rows=data_rows))
    finally:
        wb.close()

    if not sections:
        raise ParseError(path, t("parser.xlsx_empty", path=path))

    metadata: dict[str, str] = {
        "source": str(path),
        "format": "xlsx",
        "sheet_count": str(len(sections)),
    }
    for sheet_name, count in sheet_row_counts.items():
        metadata[f"row_count_{sheet_name}"] = str(count)

    return ParsedDocument(
        title=path.stem,
        sections=sections,
        tables=tables,
        metadata=metadata,
    )
